from django.conf import settings
from django.db.models.manager import Manager
from company.models import company
from django.shortcuts import render,redirect, get_object_or_404
from employee.forms import candidateProfile
from company.forms import *
from company.models import *
from .methods import *
from django.urls import reverse
from django.template.defaulttags import register
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from reportlab.pdfgen import canvas
from django.core.mail import send_mail as sendmail
from django.template.loader import render_to_string
# from weasyprint import HTML
import tempfile
from django.forms import inlineformset_factory
import openpyxl
from django.contrib.auth.decorators import login_required




@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)


@login_required(login_url='/login/2')
def home_view(request,*args):
    path = request.path
    print(path)
    if path == '/company/home/':
        guni = False
    else:
        guni = True
    current_company = get_object_or_404(company,email=request.user.email)
    profile_info = Company_profile.objects.filter(company=current_company)
    cm_name=company.objects.filter(email=request.user.email)
    email=request.user.email
    info = get_object_or_404(company, email=request.user.email)
    if guni:
        jobs = Job.objects.filter(company=info,university='Ganpat University')
    else:
        jobs = Job.objects.filter(company=info,university='Other')
    print(email)
    print(info)
    # for i in info:
    #     print(i.email)
    context={
        'info':info,
        'jobs':jobs,
        'cm_name':cm_name,      
            }
    if profile_info:
        context["profile_info"] = profile_info[0]
    if guni:
        print('guni')
        return render(request,"company/guni.html",context)
    else:
        print('comp')
        return render(request,"company/home.html",context)


@login_required(login_url='/login/2')
def create_job_view(request,*args):
    path = request.path
    print(path)
    if path == '/company/new-job/':
        guni = False
    else:
        guni = True
    current_company = get_object_or_404(company,email=request.user.email)
    # info = CompanyForm.objects.filter(company=current_company)/
    # for i in info:
    #     print(i)
    if guni:
        form = UniJobForm(request.POST or None)
    else:
        form = JobForm(request.POST or None)
    if form.is_valid():
        details = form.cleaned_data
        newId = id_generator()
        newTitle = details['job_title']
        newDescription = details['job_description']
        newPosition = details['position']
        newExp = details['experience_required']
        newVacancies = details['vacancies']
        newSalary = details['salary']
        newMaxSalary = details['maxSalary']
        newSSC = details['ssc']
        newHSC = details['hsc']
        newCGPA = details['graduation_cgpa']
        newOnline = details['aptitude_test']
        newInterview = details['video_chat']
        if guni:
            newuniversity = 'Ganpat University'
            newyear = details['pass_year']
            newlocation = details['location']
            newcampusdate = details['campus_date']
        else:
            newuniversity = 'Other'
            newyear = ''
            newlocation = details['location']
            newcampusdate = None

        if newSSC=='':
            newSSC=None
        if newHSC=='':
            newHSC=None
        if newCGPA=='':
            newCGPA=None
        newJob = Job(
            company=current_company,
            job_id=newId,
            job_title=newTitle,
            job_description=newDescription,
            salary=newSalary,
            maxSalary=newMaxSalary,
            position=newPosition,
            experience_required=newExp,
            vacancies=newVacancies,
            ssc=newSSC,
            hsc=newHSC,
            graduation_cgpa=newCGPA,
            is_active=True,
            aptitude_test = newOnline,
            video_chat = newInterview,
            university = newuniversity,
            year = newyear,
            location = newlocation,
            campus_date = newcampusdate

        )
        newJob.save()
        if guni:
            return redirect('/company/guni')
        else:
            return redirect('/company/home')
    else:
        if guni:
            form = UniJobForm(request.POST or None)
        else:
            form = JobForm(request.POST or None)       
        for field in form.errors:
            form[field].field.widget.attrs['class'] += 'error'

    context = {
        'form':form,
    }

    return render(request,"company/create_job.html", context)

@login_required(login_url='/login/2')
def profile_view(request,*args):
    cm_name=company.objects.filter(email=request.user.email)
    # info = get_object_or_404(company, email=request.user.email)
    current_company = get_object_or_404(company,email=request.user.email)
    profile_info = Company_profile.objects.filter(company=current_company)
   
    context={
        'cm_name':cm_name,
    }
    if profile_info:
        context["profile_info"] = profile_info[0]
    return render(request,"company/profile.html",context)

@login_required(login_url='/login/2')
def complete_profile_view(request,*args):
    current_company = get_object_or_404(company,email=request.user.email)
    profile_info = Company_profile.objects.filter(company=current_company)
    if not request.user.is_authenticated:
        return redirect(reverse('login'))
    form = CompanyProfile(request.POST or None)
    if request.method == 'POST':
        form = CompanyProfile(request.POST or None,request.FILES)
        if form.is_valid():
            details = form.cleaned_data
            companylink= details['companyLink']
            companyabout = details['about']
            companyemail=details['email']
            current_company = get_object_or_404(company,email=request.user.email)
            new_profile = Company_profile(
                email=companyemail,
                company=current_company,
                companyLink=companylink,
                about=companyabout,
                image = request.FILES['image'] if request.FILES.get('image') else None, 
            )
            new_profile.save()
            return redirect('/company/profile')
        else:
            form = CompanyProfile(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
        
    personal ={
    }
    if profile_info:
        personal["profile_info"] = profile_info[0]
    return render(request,"company/complete_profile.html",{'form':form,'personal':personal})

@login_required(login_url='/login/2')
def edit_update_job_view(request,*args):
    current_company = get_object_or_404(company,email=request.user.email)
    
    path = request.path
    print(path)
    if 'guni' in path:
        print('yes')
        form=Job.objects.filter(company=current_company,university='Ganpat University')
    else:
        print('no')
        form=Job.objects.filter(company=current_company,university='Other')

    
    context={
        'form':form
    }
    return render(request,"company/edit_update_job.html",context)

@login_required(login_url='/login/2')
def applications_view(request,*args):
    current_company = get_object_or_404(company,email=request.user.email)
    form=Job.objects.filter(company=current_company)
    count = {}
    for i in form:
        print(i.job_id)
        count[i.job_id] = len(job_application.objects.filter(job=i.job_id))

    context={
        'form':form,
        'count':count,
    }
    return render(request,"company/applications.html",context)

@login_required(login_url='/login/2')
def guni_applications_view(request,*args):
    current_company = get_object_or_404(company,email=request.user.email)
    form=Job.objects.filter(company=current_company)
    count = {}
    for i in form:
        print(i.job_id)
        count[i.job_id] = len(job_application.objects.filter(job=i.job_id))

    context={
        'form':form,
        'count':count,
    }
    return render(request,"company/guni_applications.html",context)

@login_required(login_url='/login/2')
def all_applications_view(request,company_code,job_id,*args):
    print(job_id)
    data=job_application.objects.filter(job=job_id)

    # job_name=Job.objects.filter(job_id=job_id)
    print(data)
    for i in data:
        print(i.company)
        print(i.candidate)
    context={
        'form':data,
        'job_id':job_id,
        # 'job_name':job_name
    }
    return render(request,"company/allapplications.html",context)

@login_required(login_url='/login/2')
def export_excel_view(request,job_id,*args):
    data=job_application.objects.filter(job=job_id)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename={date}-Applicants.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Applications'

    # Define the titles for columns
    columns = [
        'Job',
        'Position',
        'Experience',
        'Name',
        'Email',
        'Mobile',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for i in data:
        print('titlle',i.job.job_title)
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            i.job.job_title,
            i.job.position,
            i.job.experience_required,
            i.candidate.firstName + ' ' + i.candidate.lastName,
            i.candidate.email,
            i.candidate.mobile,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    res=response
    print(res)
    return response
   
@login_required(login_url='/login/2')
def ganpat_view(request,*args):
    return render(request,"company/guni.html")

@login_required(login_url='/login/2')
def company_job_edit_view(request,job_id,*args):
    print(job_id)
    current_company = get_object_or_404(company,email=request.user.email)
    displaydata=Job.objects.get(job_id=job_id)
    updatedata=Job.objects.get(job_id=job_id)
    print(updatedata.job_description)
    if request.method == "POST":
        print('post')
        form=jobeditform(request.POST or None,instance=updatedata)
        for i in form:
            print(i)
        #error here
        if form.is_valid():
            print('valid')
            form.save()
            return redirect("/company/home/")
        else:
            print('error')
    context={
        'editdata':displaydata
    }
    return render(request,"company/jobedit.html",context)

@login_required(login_url='/login/2')
def company_schedule_view(request,job_id,*args):
    path = request.path
    jobobj = Job.objects.filter(job_id=job_id)
    print(path)
    if path == '/company/applications/'+str(job_id):
        guni = False
    else:
        guni = True
    current_company = get_object_or_404(company,email=request.user.email)
    # info = CompanyForm.objects.filter(company=current_company)/
    # for i in info:
    #     print(i)
    if guni:
        form = schedule(request.POST or None)
    else:
        form = schedule(request.POST or None)
    if form.is_valid():
        details = form.cleaned_data
        if guni:
            newuniversity = 'Ganpat University'
            newdate = details['test_date']
            newtime = details['test_time']
          
        else:
            newuniversity = 'Other'
            newdate = details['test_date']
            newtime = details['test_time']

        newJob = Test_Schedule(
            job = jobobj[0],
            test_date=newdate,
            test_time =newtime,
            university=newuniversity

        )
        newJob.save()
        if guni:
            return redirect('/company/guni')
        else:
            return redirect('/company/home')
    else:
        if guni:
            form = schedule(request.POST or None)
        else:
            form = schedule(request.POST or None)       
        for field in form.errors:
            form[field].field.widget.attrs['class'] += 'error'

    scheduled = Test_Schedule.objects.filter(job=jobobj[0])

    context = {
        'form':form,
        'scheduled':scheduled,
        'job_id':job_id,
    }
    return render(request,"company/schedule.html",context)

# def online_schedule_view(request,job_id,*args):
#     path = request.path
#     jobobj = Job.objects.filter(job_id=job_id)
#     print(path)
#     if path == '/company/applications/'+str(job_id)+'/online':
#         guni = False
#     else:
#         guni = True
#     current_company = get_object_or_404(company,email=request.user.email)
#     # info = CompanyForm.objects.filter(company=current_company)/
#     # for i in info:
#     #     print(i)
#     if guni:
#         form = schedule(request.POST or None)
#     else:
#         form = schedule(request.POST or None)
#     if form.is_valid():
#         details = form.cleaned_data
#         if guni:
#             newuniversity = 'Ganpat University'
#             newdate = details['test_date']
#             newtime = details['test_time']
          
#         else:
#             newuniversity = 'Other'
#             newdate = details['test_date']
#             newtime = details['test_time']

#         newJob = Test_Schedule(
#             job = jobobj[0],
#             test_date=newdate,
#             test_time =newtime,
#             university=newuniversity

#         )
#         newJob.save()
#         if guni:
#             return redirect('/company/guni')
#         else:
#             return redirect('/company/home')
#     else:
#         if guni:
#             form = schedule(request.POST or None)
#         else:
#             form = schedule(request.POST or None)       
#         for field in form.errors:
#             form[field].field.widget.attrs['class'] += 'error'

#     scheduled = Test_Schedule.objects.filter(job=jobobj[0])

#     context = {
#         'form':form,
#         'scheduled':scheduled,
#         'job_id':job_id,
#     }
#     return render(request,"company/online_schedule.html",context)

@login_required(login_url='/login/2')
def recruitment_view(request,*args):
    current_company = get_object_or_404(company,email=request.user.email)
    
    path =request.path
    print(path)
    st={}
    create_disabled = []
    results_disabled = []
    questions_disabled = []
    selected_disabled = []

    if path =='/company/recruitment/guni':
        form=Job.objects.filter(company=current_company,university='Ganpat University')
    
        for i in form:
            schedule_time = Test_Schedule.objects.filter(job=i)
            if schedule_time:
                st[i.job_id] = str(schedule_time[0].test_time.strftime("%I:%M %p"))+' on '+str(schedule_time[0].test_date.strftime("%d/%m/%Y"))
            else:
                st[i.job_id] = 'Not Scheduled'
    
        context = {
        'form':form,
        'schedule':st
    }
        return render(request,"company/guni_recruitment.html",context)
    else:
        form=Job.objects.filter(company=current_company,university='Other')
        for i in form:
            schedule_time = Test_Schedule.objects.filter(job=i)
            if schedule_time:
                st[i.job_id] = str(schedule_time[0].test_time.strftime("%I:%M %p"))+' on '+str(schedule_time[0].test_date.strftime("%d/%m/%Y"))
            else:
                st[i.job_id] = 'Not Scheduled'
                create_disabled.append(i.job_id)
            

            quiz_obj = quiz.objects.filter(job=i)

            if quiz_obj:
                create_disabled.append(i.job_id)
                results = Marks_Of_Candidate.objects.filter(quiz=quiz_obj[0])
                if not results:
                    results_disabled.append(i.job_id)
                else:
                    selected_obj = selected_candidates.objects.filter(job=i.job_id)
                    if selected_obj:
                        selected_disabled.append(i.job_id)
            else:
                results_disabled.append(i.job_id)
                questions_disabled.append(i.job_id)


        print(st)
        print(results_disabled)
        context = {
        'form':form,
        'schedule':st,
        'results_disabled':results_disabled,
        'questions_disabled':questions_disabled,
        'create_disabled':create_disabled,
        'selected_disabled':selected_disabled
    }
  
        return render(request,"company/recruitment.html",context)

@login_required(login_url='/login/2')
def create_quiz_view(request,job_id,*args):
    print(job_id)
    jobobj = Job.objects.filter(job_id=job_id)
    current_company = get_object_or_404(company,email=request.user.email)
    scheduleobj = Test_Schedule.objects.filter(job=jobobj[0])
    print(scheduleobj)
    form = QuizForm(request.POST or None)
    if request.method == 'POST':
        form = QuizForm(request.POST or None)
        if form.is_valid():
            details = form.cleaned_data
            newname = details['name']
            newnumber = details['number_of_questions']
            newtime = details['time']
            new_quiz = quiz(
                    job=jobobj[0],
                    company=current_company,
                    schedule=scheduleobj[0],
                    name=newname,
                    number_of_questions=newnumber,
                    time = newtime,
                )
            new_quiz.save()
            return redirect('/company/recruitment')

    context = {
        'form':form,
        
    }
    return render(request,"company/create_quiz.html",context)

@login_required(login_url='/login/2')
def create_question_view(request,job_id,*args):
    print(job_id)
    jobobj = Job.objects.filter(job_id=job_id)
    current_company = get_object_or_404(company,email=request.user.email)
    quizobj = quiz.objects.filter(job=jobobj[0])
    form = QuestionForm(request.POST or None)
    if request.method == 'POST':
        form = QuestionForm(request.POST or None)
        if form.is_valid():
            details = form.cleaned_data
            newquestion= details['question']
            new_question = Questions(
                    question_num = id_generator(),
                    question=newquestion,
                    quiz=quizobj[0]
                )
            new_question.save()
            return redirect('/company/recruitment/question/'+str(job_id))

    ques = Questions.objects.filter(quiz=quizobj[0])
    print('q',ques)
    for i in ques:
        print(i.question)
    print(ques.count())

    # option_data=Option.objects.filter(question=ques[0])
    added_options = {}
    for i in ques:
        # print(i.company.company_name)
        # print(i.company)
        try:
            option_info=Option.objects.filter(question=i)
        except:
            option_info=None
        if option_info:
            added_options[i.question_num] = 1
    print(added_options)
    context = {
        'form':form,
        'ques':ques,
        'quiz':job_id,
        'no':str(ques.count()),
        'addedoptions':added_options,
        'quizobj':quizobj,
    }
    return render(request,"company/create_question.html",context)

@login_required(login_url='/login/2')
def create_option_view(request,job_id,question_num,*args):
    print(job_id)
    print(question_num)
    print(request.path)
    jobobj = Job.objects.filter(job_id=job_id)
    quizobj = quiz.objects.filter(job=jobobj[0])
    ques = Questions.objects.filter(quiz=quizobj[0],question_num=question_num)

    QuestionFormSet = inlineformset_factory(Questions, Option, fields=('content','correct', 'question'), extra=5)
    if request.method == "POST":
        print('post')
        formset = QuestionFormSet(request.POST, instance=ques[0])
        if formset.is_valid():
            formset.save()
            return redirect('/company/recruitment/question/'+str(job_id))
    else:
        print('error')
        formset=QuestionFormSet(instance=ques[0])

    
    context={
        'formset':formset,
        'question':ques[0],
        'job_id':job_id,
        'question_num':question_num,

    }
    return render(request,"company/add_options.html",context)

@login_required(login_url='/login/2')
def company_result_view(request,job_id,score,*args):
    print(job_id)
    print(score)
    quizname = quiz.objects.filter(job=job_id)

    if quizname:        
        test_result = Marks_Of_Candidate.objects.filter(quiz=quizname[0],score__gte=score,score__lte=score+9)
    else:
        test_result = None

    print(quizname,test_result)
    context={
        'quizname':quizname,
        'result':test_result,
        'job_id':job_id,
        'score':score,
    }
    return render(request,"company/result_view.html",context)

@login_required(login_url='/login/2')
def export_result_excel_view(request,job_id,score,*args):
    quizname=quiz.objects.filter(job=job_id)
    data = Marks_Of_Candidate.objects.filter(quiz=quizname[0],score__gte=score,score__lte=score+str(9))
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename={date}-Applicants.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Candidate Result'

    # Define the titles for columns
    columns = [
        'Job',
        'Position',
        'Quiz',
        'Name',
        'Email',
        'Mobile',
        'Score',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for i in data:
        print('title',i.quiz.job.job_title)
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            i.quiz.job.job_title,
            i.quiz.job.position,
            i.quiz.name,
            i.candidate.firstName + ' ' + i.candidate.lastName,
            i.candidate.email,
            i.candidate.mobile,
            i.score,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    res=response
    print(res)
    return response

@login_required(login_url='/login/2')
def online_schedule_view(request,job_id,*args):
    print(job_id)
    data = Job.objects.filter(job_id=job_id)
    form = onlineSchedule(request.POST or None)
    if request.method == 'POST':
        print('post')
        form = onlineSchedule(request.POST or None,request.FILES)
        if form.is_valid():
            print('valid')
            new_schedule= online_schedule(
                job = data[0],
                candidate_sheet=request.FILES['candidate_sheet'] if request.FILES.get('candidate_sheet') else print('not found'),
                )
            new_schedule.save()
            return redirect('/company/recruitment/result/'+str(job_id)+'/'+str(50)+'/process')
    context={
        'form':form,
    }
    return render(request,"company/online_schedule.html",context)


@login_required(login_url='/login/2')
def upload_selected_candidates(request,job_id,*args):
    print(job_id)
    data = Job.objects.filter(job_id=job_id)
    form = selectedCandidates(request.POST or None)
    if request.method == 'POST':
        print('post')
        form = selectedCandidates(request.POST or None,request.FILES)
        if form.is_valid():
            print('valid')
            new_selected= selected_candidates(
                job = data[0],
                selected_sheet=request.FILES['selected_sheet'] if request.FILES.get('selected_sheet') else print('not found'),
                )
            new_selected.save()
            return redirect('/company/recruitment/result/'+str(job_id)+'/selected-process')
    context={
        'form':form,
    }
    return render(request,"company/selected.html",context)
    
@login_required(login_url='/login/2')
def online_schedule_process_view(request,job_id,score,*args):
    print(score)
    jobobj = Job.objects.filter(job_id=job_id)
    data = online_schedule.objects.filter(job=jobobj[0])

    for i in data:
        print(i.candidate_sheet)
        wb = openpyxl.load_workbook(i.candidate_sheet)
        print(wb.sheetnames)
        worksheet = wb["Candidate Result"]
        print(worksheet)
    excel_data = list()

    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    print(excel_data)


    dict={}
   
    email_from = settings.EMAIL_HOST_USER
    
    for  j in excel_data[1:]:
            dict[j[4]] = [j[9], j[7], j[8]]
    print(dict)

    for i in dict:
        print('a',i)
        print(dict[i])
        recipient_list = [i]
        sendmail("message","Joining Link: "+dict[i][0]+"\nDate: "+dict[i][1]+"\nTime: "+dict[i][2],email_from,recipient_list)


    return render(request,"company/online_schedule_process.html")

@login_required(login_url='/login/2')
def selected_candidates_process(request,job_id,*args):
    
    jobobj = Job.objects.filter(job_id=job_id)
    data = selected_candidates.objects.filter(job=jobobj[0])
    print(data)
    for i in data:
        print(i.selected_sheet)
        wb = openpyxl.load_workbook(i.selected_sheet)
        print(wb.sheetnames)
        worksheet = wb["Sheet1"]
        print(worksheet)
    excel_data = list()

    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    print(excel_data)
    email=[]
    print(jobobj[0].company.company_name)

    dict={}
    for i in (excel_data):
        email.append(i[3])
    email_from = settings.EMAIL_HOST_USER
  
    for  j in excel_data[1:]:
       
        dict[j[3]] = [j[1], j[2]]

    print(dict)

    for i in dict:
        print('a',i)
        print(dict[i])
        recipient_list = [i]
        sendmail("Confirmation","Greetings from "+jobobj[0].company.company_name+"\nDear "+dict[i][1]+"\nYou have been selected for the role of "+dict[i][0],email_from,recipient_list)
    return render(request,"company/online_schedule_process.html")

@login_required(login_url='/login/2')
def candidates_selected_view(request,job_id,*args):
    print(job_id)

    context ={
        'job_id':job_id,
    }
    return render(request,"company/selected_view.html",context)

    

    
