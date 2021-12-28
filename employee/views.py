from datetime import date
from datetime import timedelta
from django.contrib.admin import options
from django.db import reset_queries
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from employee.forms import *
from employee.models import *
from company.forms import *
from company.models import *
import common
from django.http import JsonResponse
import openpyxl


from django.template.defaulttags import register
@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)

@login_required(login_url='/login/1')
def home_view(request,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    profile_info = Candidate_profile.objects.filter(candidate=current_candidate)
    ssc_data = ssc.objects.filter(candidate=current_candidate)
    hsc_data = hsc.objects.filter(candidate=current_candidate)
    graduation_data = graduation.objects.filter(candidate=current_candidate)
    experience_data = experience.objects.filter(candidate=current_candidate)

    candidate_name = Candidate_profile.objects.filter(candidate=current_candidate)

    profile_empty = False
    if not (profile_info and ssc_data and hsc_data and graduation_data):
        profile_empty = True
        context = {
            'profile_empty':profile_empty,
        }
        return render(request,"employee/home.html",context)

    l=[]
    if profile_info:
        l.append(profile_info[0].designation)
    else: 
        l.append('')

    if ssc_data:
        l.append(ssc_data[0].marks)
    else:
        l.append(0)
    
    if hsc_data:
        l.append(hsc_data[0].marks)
    else:
        l.append(0)

    if graduation_data:
        l.append(graduation_data[0].marks)
    else:
        l.append(0)

    print(l)


    c= experience_data.count()
    exp=[]
    for i in range(0,c):
        exp.append(experience_data[i])
    exp.reverse()

    total_exp = 0

    if exp:
        minyear = int(exp[0].startyear)
        maxyear = int(exp[0].endyear)
        for i in exp:
            if int(i.startyear) < minyear:
                minyear = int(i.startyear)
            if int(i.endyear) > maxyear:
                maxyear = int(i.endyear)

        total_exp = maxyear - minyear
    
    current_date = date.today()

    # cm_name=company.objects.filter(email=request.user.email)
    # email=request.user.email
    candidate_uni = current_candidate.university
    print(candidate_uni)

    if candidate_uni == 'Ganpat University' and graduation_data:
            passyear = graduation_data[0].passyear
    else:
        passyear = ''
    
    data=0
    if l[0]=='Fresher':
        info=Job.objects.filter(experience_required=0,ssc__lte=l[1] ,hsc__lte=l[2],graduation_cgpa__lte=l[3],university=candidate_uni,year=passyear,campus_date__gte=current_date) 
    else:
        info=Job.objects.filter(experience_required__lte=total_exp,ssc__lte=l[1] ,hsc__lte=l[2],graduation_cgpa__lte=l[3],university=candidate_uni,year=passyear,campus_date__gte=current_date)
    
    if l[0]=='Fresher':
        info2=Job.objects.filter(experience_required=0,ssc=None ,hsc=None,graduation_cgpa=None,university=candidate_uni,year=passyear,campus_date__gte=current_date)
    else:
        info2=Job.objects.filter(experience_required__lte=total_exp,ssc=None ,hsc=None,graduation_cgpa=None,university=candidate_uni,year=passyear,campus_date__gte=current_date)

    if candidate_uni == 'Other':
        
        if l[0]=='Fresher':
            info=Job.objects.filter(experience_required=0,ssc__lte=l[1] ,hsc__lte=l[2],graduation_cgpa__lte=l[3],university=candidate_uni) 
        else:
            info=Job.objects.filter(experience_required__lte=total_exp,ssc__lte=l[1] ,hsc__lte=l[2],graduation_cgpa__lte=l[3],university=candidate_uni)
    
        if l[0]=='Fresher':
            info2=Job.objects.filter(experience_required=0,ssc=None ,hsc=None,graduation_cgpa=None,university=candidate_uni)
        else:
            info2=Job.objects.filter(experience_required__lte=total_exp,ssc=None ,hsc=None,graduation_cgpa=None,university=candidate_uni)

    
    for i in info2:
        print('b',i)

    info_data=info|info2

    print(info_data)

    no_jobs = False
    if len(info_data)==0:
        no_jobs=True

    cl={}
    applied_jobs = {}
    for i in info_data:
        try:
            profile_info=Company_profile.objects.get(company=i.company)
        except:
            profile_info=None
        cl[i.company.company_name]=profile_info
        applied = job_application.objects.filter(candidate=current_candidate,job=i)
        if applied:
            applied_jobs[i.job_id] = 1
            
    print(cl)
    
    
    context={
        'info':reversed(info_data),
        'applied_jobs':applied_jobs,
        'candidate_name':candidate_name,
        'no_jobs':no_jobs,
           }
    if profile_info:
        context['profile_info']=cl

    return render(request,"employee/home.html",context)

@login_required(login_url='/login/1')
def interest_view(request,*args):
    current_candidate= get_object_or_404(candidate,email=request.user.email)
    data = job_application.objects.filter(candidate=current_candidate)
    results_blocked = []
    for i in data:
        print(i.job.job_id)
        schedules = online_schedule.objects.filter(job=i.job)
        if not schedules:
            results_blocked.append(i.job.job_id)

    print(results_blocked)
    context={
        'data':reversed(data),
        'results_blocked':results_blocked,
    }

    return render(request,"employee/applications.html",context)


@login_required(login_url='/login/1')
def recruitment_view(request,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    data =job_application.objects.filter(candidate=current_candidate)
    tests = {}
    quizes= {}
    quiz_available = []
    for i in data:
        test = Test_Schedule.objects.filter(job=i.job) 
        tests[i.job] =  test
        if test:
            date = test[0].test_date
            time = test[0].test_time
            print(datetime.today().date())
            if datetime.today().date() == date:
                t = datetime.now().time()
                delta = timedelta(minutes=15)
            
                if datetime.now().time() >= time and datetime.now().time() <= (datetime.combine(datetime.today().date(),t) + delta).time():
                    quiz_obj = quiz.objects.filter(job=i.job)
                    if quiz_obj:
                        questions = Questions.objects.filter(quiz=quiz_obj[0])
                        if questions:
                            attempted = Marks_Of_Candidate.objects.filter(quiz=quiz_obj[0],candidate=current_candidate)
                            if not attempted:
                                quiz_available.append(i.job.job_id)

        
    
        quizes[i.job] = quiz.objects.filter(job=i.job)
    # print('a',tests,quizes)
    
    print(quiz_available)
    # print(tests)
    context ={
        'tests':tests,
        'data':data,
        'quiz':quizes,
        'quiz_available':quiz_available,
    }

    return render(request,"employee/recruitment.html",context)

@login_required(login_url='/login/1')
def profile_view(request,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    profile_info = Candidate_profile.objects.filter(candidate=current_candidate)
    ssc_data = ssc.objects.filter(candidate=current_candidate)
    hsc_data = hsc.objects.filter(candidate=current_candidate)
    graduation_data = graduation.objects.filter(candidate=current_candidate)
    experience_data = experience.objects.filter(candidate=current_candidate)
    social_data = social.objects.filter(candidate=current_candidate)
    candidate_data =candidate.objects.filter(email=request.user.email)
    # pic=Candidate_profile.objects.filter('image',candidate=current_candidate)
    # print(profile_info['image'])

    c= experience_data.count()
    l=[]
    for i in range(0,c):
        l.append(experience_data[i])
    l.reverse()
    
    context = {}

    if l:
        minyear = int(l[0].startyear)
        maxyear = int(l[0].endyear)
        for i in l:
            if int(i.startyear) < minyear:
                minyear = int(i.startyear)
            if int(i.endyear) > maxyear:
                maxyear = int(i.endyear)

        total_exp = maxyear - minyear
        context["total_exp"] = total_exp

    print(l)
    
    # print(exp)
    # for i in exp:
    #     print(i)
    
    if candidate_data:
        context["candidate_data"] = candidate_data[0]
    if profile_info:
        context["profile_info"] = profile_info[0]
    if ssc_data:
        context["ssc_data"] = ssc_data[0]
    if hsc_data:
        context["hsc_data"] = hsc_data[0]    
    if graduation_data:
        context["graduation_data"] = graduation_data[0]
    # for i in range(0,c):   
    if experience_data:
        context["experience_data"] = l
    if social_data:
        context["social_data"] = social_data[0]

    # print(context)
    return render(request,"employee/profile.html",context)

@login_required(login_url='/login/1')
def complete_profile_view(request,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    profile_info = Candidate_profile.objects.filter(candidate=current_candidate)
    if not request.user.is_authenticated:
        return redirect(reverse('login'))

    form = candidateProfile(request.POST or None)
    if request.method == 'POST':
        form = candidateProfile(request.POST or None,request.FILES)
        if form.is_valid():
            details = form.cleaned_data
            newFname = details['firstName']
            newMname = details['middleName']
            newLname = details['lastName']
            newBio = details['about']
            newBirthdate = details['date']
            newGender = details['gender']
            newDesignation = details['designation']
            current_candidate = get_object_or_404(candidate,email=request.user.email)
            new_profile = Candidate_profile(
                candidate=current_candidate,
                firstName=newFname,
                middleName=newMname,
                lastName=newLname,
                about=newBio,
                date=newBirthdate,
                gender=newGender,
                designation=newDesignation,
                image = request.FILES['image'] if request.FILES.get('image') else None, 
            )
            new_profile.save()
            return redirect('/candidate/profile')
        else:
            form = candidateProfile(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
        
    personal ={
    }
    if profile_info:
        personal["profile_info"] = profile_info[0]
    return render(request,"employee/complete_profile.html",{'form':form,'personal':personal})

@login_required(login_url='/login/1')
def education_profile_view(request,*args):
    return render(request,"employee/education_profile.html")

@login_required(login_url='/login/1')
def ssc_profile_view(request,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    ssc_data = ssc.objects.filter(candidate=current_candidate)
    re=request.path.split('/')
    form=sscProfile(request.POST or None)
    if request.method == 'POST':
        form=sscProfile(request.POST or None)
        if form.is_valid():
            details = form.cleaned_data
            start_year = details['startyear']
            end_year = details['passyear']
            institute = details['instituteName']
            percentage = float(details['marks'])
            percentage = "{:.2f}".format(percentage)
            current_candidate = get_object_or_404(candidate,email=request.user.email)
            new_ssc = ssc(
                candidate=current_candidate,
                startyear=start_year,
                passyear=end_year,
                instituteName=institute,
                marks=percentage,
                ssc_marksheet=request.FILES['ssc_marksheet'] if request.FILES.get('ssc_marksheet') else print('not found'), 
            )
            
            new_ssc.save()
            return redirect('/candidate/profile')

        else:
            form = sscProfile(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
    context ={
        'form':form,
        're':(re[4].upper())
    }
    if ssc_data:
        context['ssc_data']= ssc_data[0]
    return render(request,"employee/education_submit.html",context)

@login_required(login_url='/login/1')
def hsc_profile_view(request,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    hsc_data = hsc.objects.filter(candidate=current_candidate)
    form=hscProfile()
    re=request.path.split('/')
    form=hscProfile(request.POST or None)
    if request.method == 'POST':
        form=hscProfile(request.POST or None)
        if form.is_valid():
            details = form.cleaned_data
            category = details['categories']
            start_year = details['startyear']
            end_year = details['passyear']
            institute = details['instituteName']
            percentage = float(details['marks'])
            percentage = "{:.2f}".format(percentage)
            current_candidate = get_object_or_404(candidate,email=request.user.email)
            new_hsc = hsc(
                candidate=current_candidate,
                category=category,
                startyear=start_year,
                passyear=end_year,
                instituteName=institute,
                marks=percentage,
                hsc_marksheet=request.FILES['hsc_marksheet'] if request.FILES.get('hsc_marksheet') else print('not found'), 

            )
            new_hsc.save()
            return redirect('/candidate/profile')
            
        else:
            form = hscProfile(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
    context={
        'form':form,
        're':re[4].upper()
    }
    if hsc_data:
        context['hsc_data']=hsc_data[0]
    return render(request,"employee/education_submit.html",context)

@login_required(login_url='/login/1')
def graduation_profile_view(request,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    graduation_data = graduation.objects.filter(candidate=current_candidate)
    form=graduationProfile()
    re=request.path.split('/')
    form=graduationProfile(request.POST or None)
    if request.method == 'POST':
        form=graduationProfile(request.POST or None)
        if form.is_valid():
            details = form.cleaned_data
            category = details['categories']
            start_year = details['startyear']
            end_year = details['passyear']
            institute = details['instituteName']
            percentage = float(details['marks'])
            percentage = "{:.2f}".format(percentage)
            current_candidate = get_object_or_404(candidate,email=request.user.email)
            new_grad = graduation(
                candidate=current_candidate,
                category=category,
                startyear=start_year,
                passyear=end_year,
                instituteName=institute,
                marks=percentage,
                graduation_marksheet=request.FILES['graduation_marksheet'] if request.FILES.get('graduation_marksheet') else print('file not found'), 

            )
            new_grad.save()
            return redirect('/candidate/profile')
            
        else:
            form = graduationProfile(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
    context={
        'form':form,
        're':re[4].upper()
    }
    if graduation_data:
        context['graduation_data']=graduation_data[0]
    return render(request,"employee/education_submit.html",context)

@login_required(login_url='/login/1')
def experience_profile_view(request,*args):
    form=experienceProfile()
    re=request.path.split('/')
    form=experienceProfile(request.POST or None)
    if request.method == 'POST':
        form=experienceProfile(request.POST or None)
        if form.is_valid():
            details = form.cleaned_data
            start_year = details['startyear']
            end_year = details['endyear']
            company = details['companyName']
            role = details['role']
            description =details['about']
            ctc = float(details['ctc'])
            ctc = "{:.1f}".format(ctc)
            current_candidate = get_object_or_404(candidate,email=request.user.email)
            new_exp = experience(
                candidate=current_candidate,
                startyear=start_year,
                endyear=end_year,
                role=role,
                about=description,
                companyName=company,
                ctc=ctc,
                experience_file=request.FILES['experience_file'] if request.FILES.get('experience_file') else None, 

            )
            new_exp.save()
            return redirect('/candidate/profile')
            
        else:
            form = experienceProfile(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
    context={
        'form':form,
        're':re[4].upper()
    }
    return render(request,"employee/education_submit.html",context)

@login_required(login_url='/login/1')
def other_profile_view(request,*args):
    return render(request,"employee/other_profile.html")


@login_required(login_url='/login/1')
def social_profile_view(request,*args):
    form = socialProfile()
    re=request.path.split('/')
    form=socialProfile(request.POST or None)
    if request.method == 'POST':
        form=socialProfile(request.POST or None)
        if form.is_valid():
            details = form.cleaned_data
            city = details['city']
            linkedin = details['linkedin']
            github = details['github']
            current_candidate = get_object_or_404(candidate,email=request.user.email)
            new_social = social(
                candidate=current_candidate,
                city=city,
                linkedin=linkedin,
                github=github
            )
            new_social.save()
            return redirect('/candidate/profile')
            
        else:
            form = socialProfile(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
    context = {
        'form':form,
         're':re[4].upper()
    }
    return render(request,"employee/social_profile.html",context)

@login_required(login_url='/login/1')
def job_application_view(request,company_code,job_id,*args):
    current_candidate = get_object_or_404(candidate,email=request.user.email)
    profile_info = Candidate_profile.objects.get(candidate=current_candidate)
    experience_data = experience.objects.filter(candidate=current_candidate)
    c= experience_data.count()
    l=[]
    for i in range(0,c):
        l.append(experience_data[i])
    l.reverse()
    

    if l:
        minyear = int(l[0].startyear)
        maxyear = int(l[0].endyear)
        for i in l:
            if int(i.startyear) < minyear:
                minyear = int(i.startyear)
            if int(i.endyear) > maxyear:
                maxyear = int(i.endyear)

        total_exp = maxyear - minyear
       

    print(l)
    
    # data_dict={'firstName':profile_info.firstName,'lastName':profile_info.lastName,'email':request.user.email,
    #             'mobile':current_candidate.mobile,'total_experience':total_exp}

    # print(data_dict)
    form = JobApplicationForm(request.POST or None, request.FILES)

    for i in form:
        print(i)
    if request.method == 'POST':
        form = JobApplicationForm(request.POST or None,request.FILES)
        if form.is_valid():
            details = form.cleaned_data
            company_data=company.objects.get(account_id=company_code)
            job_data=Job.objects.get(job_id=job_id)
            new_application = job_application(
                candidate=current_candidate,
                company=company_data,
                job=job_data,
                resume_file= request.FILES['resume_file'] if request.FILES.get('resume_file') else None, 
            )
            new_application.save()
            return redirect('/candidate/home')
        else:
            form = JobApplicationForm(request.POST or None)
            for field in form.errors:
                    print(field)
                    form[field].field.widget.attrs['class'] += 'error'
        
  
    return render(request,"employee/job_application_form.html",{'form':form})


@login_required(login_url='/login/1')
def candidate_schedule_view(request,job,*args):
    print(job)
    current_candidate= get_object_or_404(candidate,email=request.user.email)
    data = job_application.objects.filter(candidate=current_candidate)
    print(data)
    for i in data:
        print(i.job.job_id)
        if job == i.job.job_id:
            print('yes')
            schedule_time= Test_Schedule.objects.filter(job=i.job)
            quizs= quiz.objects.filter(job=i.job)
            try:
                candidate_marks = Marks_Of_Candidate.objects.get(quiz=quizs[0],candidate=current_candidate)
                candidate_marks=str(candidate_marks.score) + ' %'
            except:
                candidate_marks = 'Not Declared'

                
            print(schedule_time)
    print('a',schedule_time)

    context = {
        'schedule_time':schedule_time,
        'data':data,
        'marks':candidate_marks,
    }

    return render(request,"employee/schedule.html",context)

@login_required(login_url='/login/1')
def candidate_quiz_view(request,job,*args):
    print(job)
    current_candidate= get_object_or_404(candidate,email=request.user.email)
    data = job_application.objects.filter(candidate=current_candidate)
    print(data)
    attempt_block = False
    for i in data:
        print(i.job.job_id)
        if job == i.job.job_id:
            quizs= quiz.objects.filter(job=i.job)
            if quizs:
                attempted = Marks_Of_Candidate.objects.filter(quiz=quizs[0],candidate=current_candidate)
                if attempted:
                    attempt_block = True
            print(quizs)
    print('a',quizs)
    for i in quizs:
        print(i)
    context = {
        'quiz':quizs,
        'data':data,
        'attempt_block':attempt_block,
    }
    return render(request,"employee/quiz.html",context)

@login_required(login_url='/login/1')
def quiz_start(request,job,time,name,*args):
    print(job)
    print(time)
    current_candidate= get_object_or_404(candidate,email=request.user.email)
    data = job_application.objects.filter(candidate=current_candidate)
    print(data)
    for i in data:
        print(i.job.job_id)
        if job == i.job.job_id:
            print('yes')
            quizs= quiz.objects.filter(job=i.job)
            print(quizs)
    print('a',quizs)
    for i in quizs:
        print(i)
   
    print(job,name)
   
    que =  Questions.objects.filter(quiz=quizs[0])
    print(len(que))
    total_que=len(que)
 
    options_data={}
    for i in que:
       
        option_data = Option.objects.filter(question=i)
        options_data[i.question]=option_data
    
    print('ac',options_data)
    print('ad',que)

    if request.method=="POST":
        answers = list(request.POST.items())
        print(answers)
        c=0
        for i in range(1,len(answers)):
            quest = answers[i]
            question = quest[0]
            ans = quest[1]
            queobj = Questions.objects.get(question=question)
            ansobj = Option.objects.get(question=queobj,content=ans)
            print('queobj',queobj)
            print('ansobj',ansobj)

            if ansobj.correct==True:
                c+=1
        print(c)
        print(total_que)
        score = (c/total_que)*100
        print(score)
        new_marks = Marks_Of_Candidate(quiz=quizs[0],candidate=current_candidate,score=score)
        new_marks.save()
        job_id = str(job)
        # url = 'candidate/applications/schedule-view/'+
        return redirect('/candidate/applications/schedule-view/'+job_id)
            


    context = {
        'quiz':quizs,
        'data':data,
        'question':que,
        'options_data':options_data,
    }
    return render(request,"employee/quizstarted.html",context)

@login_required(login_url='/login/1')
def candidates_result_view(request,job_id,*args):
    candidate_email=request.user.email
    print(job_id)
    quizname = quiz.objects.filter(job=job_id)

    jobobj = Job.objects.filter(job_id=job_id)
    data = online_schedule.objects.filter(job=jobobj[0])
    
    print(data)

    for i in data:
        print(i.candidate_sheet)
        wb = openpyxl.load_workbook(i.candidate_sheet)
        print(wb.sheetnames)
        worksheet = wb["Candidate Result"]
        print(worksheet)
    excel_data = list()

    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    print(excel_data[1:])

    context={
        'quizname':quizname[0],
        'data':excel_data[1],
    }
    for i in excel_data[1:]:
        if request.user.email==i[4]:
            i[7]= i[7].replace(" 00:00:00","")
            context['data']=list(i)
            break
    if not context['data']:
        context['data']='Not Found'

    return render(request,"employee/result_view.html",context)
